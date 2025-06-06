# TestY TMS - Test Management System
# Copyright (C) 2024 KNS Group LLC (YADRO)
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU Affero General Public License as published
# by the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Affero General Public License for more details.
#
# You should have received a copy of the GNU Affero General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
# Also add information on how to contact you by electronic and paper mail.
#
# If your software can interact with users remotely through a computer
# network, you should also make sure that it provides a way for users to
# get its source.  For example, if your program is a web application, its
# interface could display a "Source" link that leads users to an archive
# of the code.  There are many ways you could offer source, and different
# solutions will be better for different programs; see section 13 for the
# specific requirements.
#
# You should also get your employer (if you work as a programmer) or school,
# if any, to sign a "copyright disclaimer" for the program, if necessary.
# For more information on this, and how to apply and follow the GNU AGPL, see
# <http://www.gnu.org/licenses/>.
from django.db import transaction
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from simple_history.utils import bulk_create_with_history
from testy.tests_description.models import TestCase, TestSuite

from .exceptions import InvalidXlsx


class XlsxParser:
    def __init__(self, xlsx_file: bytes, project_id: int, test_suit_name: str):
        self.ws: Worksheet = load_workbook(xlsx_file).active
        self.project_id = project_id
        self.test_suit_name = test_suit_name

    @transaction.atomic
    def create_suites_with_cases(self):
        cases = []
        suites_counter = 0
        tmp_suite = TestSuite.objects.get(
                project_id=self.project_id,
                name=self.test_suit_name,
            )       
        for idx, row in enumerate(self.ws.iter_rows(), 1):
            try:
                number_cell, priority_cell, subject_cell, description_cell = row
            except ValueError:
                raise InvalidXlsx(
                    f'Too many values in line: expected 4, got {len(row)}',
                )

            # if not suite_cell.value and not tmp_suite:
            #     raise InvalidXlsx('Empty suite')

            # if suite_cell.value:
            #     tmp_suite = TestSuite.objects.create(
            #         project_id=self.project_id,
            #         name=suite_cell.value,
            #     )
            #     suites_counter += 1



            empty_cells = []
            if not number_cell.value:
                empty_cells.append("number_cell")
            if not priority_cell.value:
                empty_cells.append("priority_cell")
            if not subject_cell.value:
                empty_cells.append("subject_cell")
            if not description_cell.value:
                empty_cells.append("description_cell")

            if empty_cells:
                raise InvalidXlsx(f"Got empty cell(s) in line {idx}: {', '.join(empty_cells)}")

            # if not number_cell.value or not priority_cell.value or not subject_cell.value or not description_cell.value:
            #     raise InvalidXlsx(f'Got empty cell in line {idx}')

            cases.append(
                TestCase(
                    project_id=self.project_id,
                    suite = tmp_suite,
                    scenario=description_cell.value,
                    name=subject_cell.value,
                    description=number_cell.value,
                ),
            )

        bulk_create_with_history(cases, TestCase)
        return suites_counter, len(cases)
